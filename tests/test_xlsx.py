import os.path

from openpyxl import load_workbook
from tests.constants import RESOURCES_PATH

# TODO оформить в тест, добавить ассерты и использовать универсальный путь

def test_xlsx():
    xlsx_file_path = os.path.join(RESOURCES_PATH, 'file_example_XLSX_50.xlsx')
    workbook = load_workbook(xlsx_file_path)
    sheet = workbook.active
    print(sheet.cell(row=3, column=2).value)


    size = os.path.getsize(xlsx_file_path)
    assert size == 7360
    assert (sheet.cell(row=3, column=2).value) == 'Mara'

